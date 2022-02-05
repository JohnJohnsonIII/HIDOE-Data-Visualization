# import libraries
from os import write
import csv
import sys
from datetime import datetime, date
import json
import os
import openpyxl
import re
from openpyxl.utils import range_boundaries
filename = sys.argv[1]

def unmergeExcel( filename ):
    wbook=openpyxl.load_workbook(filename)
    sheet = wbook.active

    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                print('looping')
                cell.value = top_left_cell_value
    wbook.save('unmerged'+filename)

def convertExcelToCSV( filename ):
    wb_obj = openpyxl.load_workbook( filename )
    sheet = wb_obj.active
    csvdata = ''
    line = ''
    cellnum = 0
    flag = False
    twoislandsandtwocounts = False
    secondisland = ''
    namestofix = {'Inouye Elementary School': 'Daniel K Inouye Elementary School', 'Kamehameha III Elementary School': 'King Kamehameha III Elementary School', 'Kalihi-Kai Elementary School': 'Kalihi Kai Elementary School', 'DeSilva Elementary School': 'Ernest Bowen DeSilva Elementary', 'Pope Elementary School': 'Blanche Pope Elementary School', 'Kalihi-Waena Elementary School': 'Kalihi Waena Elementary School', 'Waikoloa School': 'Waikoloa Elementary & Middle School', 'Kaumualii Elementary School': 'King Kaumualii Elementary School', 'Waimea Canyon School': 'Waimea Canyon Middle School', 'Ke Kula O Ehunuikaimalino': "Ke Kula 'o 'Ehunuikaimalino", 'Kalanianaole Elementary & Inter': 'Kalanianaole Elementary & Intermediate School', 'Kau High & Pahala Elem': 'Kau High & Pahala Elementary School', 'HI School for the Deaf &': 'HI School for the Deaf & Blind', 'Complex, District or State': 'Complex, District or State Office'}#, 'Office': 'District or State Office'

    for index, row in enumerate(sheet.iter_rows()):
        for cell in row:
            datum = cell.value
            print('line: {}'.format(index))
            #First handle exceptions in this section and then we can assign last if none afterwards
            if cellnum==0:   # COMPLEX AREA
                if datum!=None:
                    complexarea = datum
            elif cellnum==1: # SCHOOL NAME - WHEN SPLIT ON TWO LINES, THE DATA IS REALLY JUST ONE LINE OF DATA. School, Blind, office, school name plus 2+ spaces followed by numbers
                if datum != None:
                    datum = str(cell.value).strip() if cell.value != None else None
                    foundreturn = datum.find('\n')
                    if foundreturn !=-1: # This is a one-off where there is only one line of data, but the school name has 'office' after a carriage return so we simply remove the \n
                        #print('found a return on line {} in position {}: {}'.format(index, foundreturn, datum))
                        schoolname = datum.replace('\n',' ')
                        #print('changed name to {}'.format(schoolname))
                    elif datum=='School' or datum=='Office' or datum=='Blind' or datum=='Complex':
                        # We need to toss this line because it is the second line and simply is a repeat of the first. We don't want to double-count cases.
                        print(datum)
                        flag = True # Throw out this line of data since it is a duplicate
                        #print( 'on line {} and row looks like ==>{}'.format(index, row))
                        #print( 'previous line looks like ==>{}'.format(lastrow))
                    elif datum.find('  ')!=-1:
                        #print('line {}: --{}--'.format(index,datum))
                        pos = datum.find('  ')
                        schoolname = datum[:pos].strip()
                        totalstaff = datum[pos:].strip() # this works because the totalstaff field comes across as a None and therefore takes the last totalstaff, which we are setting here to the correct one. Tested and worked on all.
                        #print('set totalstaff to {}'.format(totalstaff))
                    else:
                        schoolname = datum
                    schoolname = schoolname.replace(' Elem ',' Elementary ').replace(' Inter ', ' Intermediate ').replace(' Inter,',' Intermediate School,')
                    # Now we want to fix problematic names
                    if schoolname in namestofix:
                        schoolname = namestofix[ schoolname ]
                #print(schoolname)

            elif cellnum==2: # TOTAL STAFF May be Null in the case of State Office
                if datum == 'Null':
                    totalstaff = 'N/A'
                elif datum!=None:
                    totalstaff = datum
                # else we leave totalstaff set to what it was before because it is NONE                    
            elif cellnum==3: # TOTAL STUDENTS
                if datum == 'Null':
                    totalstudents = 'N/A'
                elif datum!=None:
                    totalstudents = datum
                # else we leave totalstudents set to what it was before because it is NONE                    
            elif cellnum==4: # TOTAL STAFF AND STUDENTS
                if datum == 'Null':
                    totalcombined = 'N/A'
                elif datum!=None:
                    totalcombined = datum
                # else we leave totalcombined set to what it was before because it is NONE                    
            elif cellnum==5: # DATE CASE REPORTED
                # Line 3970 has two dates separated by spaces(August 24 and August 18)
                # For some reason, cell 5 has the data for cell 5 and 6 and cell 6 still has the data for 6, so we just get rid of the data for cell 5
                # A check of values shows it is either None or a valid date
                if datum!=None and datum.find('  ')!=-1:
                    datereported = datetime.strptime( datum.split('  ')[0].strip(), '%B %d, %Y' ).strftime('%Y/%m/%d')
                elif datum!=None:
                    datereported = datetime.strptime( datum.strip(), '%B %d, %Y' ).strftime('%Y/%m/%d')
                # else we leave datereported set to what it was before because it is NONE                    
            elif cellnum==6: # LAST DATE INDIVIDUAL WAS ON A HIDOE CAMPUS Many Nulls. Replace with Unknown
                if datum == 'Null':
                    datelastoncampus = 'Unknown'
                elif datum!=None:
                    datelastoncampus = datetime.strptime( datum, '%B %d, %Y' ).strftime('%Y/%m/%d')
                # else we leave datelastoncampus set to what it was before because it is NONE                    
            elif cellnum==7: # ISLAND There are some with Null values where it deals with State Offices, so mark these as Unspecified. We have another error in merged Island cells having two islands separated by a \n.
                # In all of these cases except one, there are two lines of the data BUT the first one has both islands and the second one has None as Island.
                # In ONE case there is only one line AND the last column(the case count) has a \n in it.
                if index>=12054 and index<12057:
                    print('line: {}  island: --{}--'.format(index, datum))
                if secondisland !='':
                    island = secondisland
                    secondisland = ''
                elif datum == 'Null':
                    island = 'Unspecified'
                elif datum!=None and datum.find('\n')!=-1:
                    print('--{}-- has a return in line {}'.format(datum, index))
                    if index!=12148:
                        island = datum.split('\n')[0].strip()
                        if island == 'Null':
                            island = 'Unspecified'
                        secondisland = datum.split('\n')[1].strip()
                        if secondisland == 'Null':
                            secondisland = 'Unspecified'
                        print('line: {} island is {}'.format(index, island))
                        print('line: {} secondisland is {}'.format(index, secondisland))
                    else:
                        island = datum
                elif datum!=None:
                    island = datum
            elif cellnum==8: # CASE COUNT Checked data and everything is either an integer or None
                if datum != None and not isinstance(datum, int):
                    #print('line {}: -->{}<--'.format(index, datum))
                    twoislandsandtwocounts = True
                    casecount = datum
                    #datum = int(str(datum).strip())
                elif datum!=None:
                    casecount = datum
            cellnum +=1
        if not flag:
            if twoislandsandtwocounts:
                #print('=={}== {}  {}'.format(island, casecount, index))
                #print(island.split('\n')[0].strip())
                #print(island.split('\n')[1].strip())
                #print(casecount.split('\n')[0].strip())
                #print(casecount.split('\n')[1].strip())
                print('island is --{}--'.format(island))
                print('casecount is --{}--'.format(casecount))
                line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island, str(casecount).split('\n')[0].strip() )
                csvdata += line
                line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, secondisland, str(casecount).split('\n')[1].strip() )
                csvdata += line
                lastrow = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, secondisland, str(casecount).split('\n')[1].strip() )
                # line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[0].strip(), str(casecount).split('\n')[0].strip() )
                # csvdata += line
                # line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[1].strip(), str(casecount).split('\n')[1].strip() )
                # csvdata += line
                # lastrow = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[1].strip(), str(casecount).split('\n')[1].strip() )
                # line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[0].strip(), str(casecount).split('\n')[0].strip() )
                # csvdata += line
                # line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[1].strip(), str(casecount).split('\n')[1].strip() )
                # csvdata += line
                # lastrow = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island.split('\n')[1].strip(), str(casecount).split('\n')[1].strip() )
                twoislandsandtwocounts = False
            else:
                line = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island, casecount )
                csvdata += line
                lastrow = '"{}","{}","{}","{}","{}","{}","{}","{}","{}"\n'.format(complexarea, schoolname, totalstaff, totalstudents, totalcombined, datereported, datelastoncampus, island, casecount )
            #print('added "{}"'.format(line))
        else:
            flag = False
        # if index==12097:# or index==12147 or index==12148 or index==12149:
        #     # for j in range(index-1, index+2):
        #     print('-->{}: {}<--'.format(index, line))
        #    flag = False
        cellnum = 0
        line = ''
    return csvdata

def openAndCleanExcelFile( filename ):
    namestofix = {'Inouye Elementary School': 'Daniel K Inouye Elementary School', 'Kamehameha III Elementary School': 'King Kamehameha III Elementary School', 'Kalihi-Kai Elementary School': 'Kalihi Kai Elementary School', 'DeSilva Elementary School': 'Ernest Bowen DeSilva Elementary', 'Pope Elementary School': 'Blanche Pope Elementary School', 'Kalihi-Waena Elementary School': 'Kalihi Waena Elementary School', 'Waikoloa School': 'Waikoloa Elementary & Middle School', 'Kaumualii Elementary School': 'King Kaumualii Elementary School', 'Waimea Canyon School': 'Waimea Canyon Middle School', 'Ke Kula O Ehunuikaimalino': "Ke Kula 'o 'Ehunuikaimalino", 'Kalanianaole Elementary & Inter': 'Kalanianaole Elementary & Intermediate School', 'Kau High & Pahala Elem': 'Kau High & Pahala Elementary School', 'HI School for the Deaf &': 'HI School for the Deaf & Blind', 'Complex, District or State': 'Complex, District or State Office'}#, 'Office': 'District or State Office'
    #returns csv-formatted string
    wb_obj = openpyxl.load_workbook( filename )
    sheet = wb_obj.active
    csvdata = ''
    line = ''
    datum = ''
    count = 0
    error = False
    previous = ['','','','','','','','','']
    mergedValue = ''
    cellcount = 0
    for index, row in enumerate(sheet.iter_rows()):
        for cell in row:
            if str(cell.value).find('\n')!=-1:
                print('found a return in line {} in cell #{}.'.format(index,cellcount))
            datum = cell.value
            if datum==None:
                print('datum is None and column is {} on row {}'.format(count,index))
                if count==7 and mergedValue!='':
                    datum = mergedValue
                elif count!=8:
                    print('setting datum to {}'.format(previous[count]))
                    datum = previous[count]
                else:
                    print('problematic line with line {}'.format(line))
                    error = True
            elif count==7 and datum=='Null':
                datum = 'Unspecified'
            elif count==1 and ( datum=='School' or datum=='Blind' or datum=='Office'):
                datum = previous[count]
            elif count==5 or count==6:
                print('datum a date so it is {}'.format(datum))
                if datum.find('  ')!=-1:
                    print(line+' is a problem.')
                    print('count is {}'.format(count))
                    part1, part2 = datum.split('  ',1)
                    part1 = part1.strip()
                    part2 = part2.strip()
                    datum = datetime.strptime( part1, '%B %d, %Y' ).strftime('%Y/%m/%d') if part1!=None and part1!='' and part1.upper()!='NULL' else ''
                    mergedValue = datetime.strptime( part2, '%B %d, %Y' ).strftime('%Y/%m/%d') if part2!=None and part2!='' and part2.upper()!='NULL' else ''
                    print( 'broken into -->{}<-- and -->{}<--'.format(datum,mergedValue))
                    previous[count] = datum
                else:
                    datum = datetime.strptime( datum, '%B %d, %Y' ).strftime('%Y/%m/%d') if datum!=None and datum!='' and datum.upper()!='NULL' else ''
                    previous[count] = datum
            else:
                previous[count] = datum
            if count==1 and datum!=None: # we need to fix the case where district office breaks across two lines
                if '\n' in datum:
                    datum = datum.replace('\n',' ')
                # now we clean up the cases where the pdf conversion combined the school and next field over
                match = re.search(r'[\w\s&\s]+\s\s+[0-9]*',datum)
                if match != None:
                    match = re.search(r'[a-zA-Z\s&\s]+',match.group())
                    datum = match.group().strip()
                # now we clean up abbreviations to make it look nicer
                datum = datum.replace(' Elem ',' Elementary ').replace(' Inter ', ' Intermediate ').replace(' Inter,',' Intermediate School,')
                # Now we want to fix problematic names
                if datum in namestofix:
                    datum = namestofix[ datum ]
            line = '{}"{}",'.format(line,datum)
            print('line is {}'.format(line))
            count += 1
            cellcount +=1
        cellcount = 0
        line = line[:-1]
        mergedValue = ''
        # clean up case where pdf to xlsx conversion split on island
        if line=='"State or District Office","Complex, District or State Office","Null","Null","Null","08/09/2021","08/06/2021","Hawaii\nOahu","1\n1"':
            line = '"State or District Office","Complex, District or State Office","Null","Null","Null","08/09/2021","08/06/2021","Hawaii","1"\n"State or District Office","Complex, District or State Office","Null","Null","Null","08/09/2021","08/06/2021","Oahu","1"'
            print('fixed it to {}'.format(line))
        # clean up cases where school name broke across two lines
        if line != '"None","School","None","None","None","None","None","None","None"' and line!='"None","Blind","None","None","None","None","None","None","None"' and not error:
            line += '\n'
            csvdata += line
        print('\n{}\n'.format(line))
        line = ''
        count = 0
        error = False
    return csvdata

def importJsonFromCsv( filename ):
    with open( filename ) as csvfile:
        jsondata = list( csv.DictReader( csvfile ) )
    return jsondata

def writeJSON( data, filename ):
    with open( filename, 'w') as fout:
        fout.write( json.dumps(data, indent=2, default=str) )

def differentiateOffices( jsondata ):
    for data in jsondata:
        if data["School Name"]=='Complex, District or State Office':
            data["School Name"] = data["Island"]+' '+data["School Name"] if data["Island"] != 'Null' else 'Unspecified' + ' ' + data["School Name"]
    return jsondata


headers = 'Complex Area,School Name,Total Staff,Total Students,Total Staff & Students,Date Case Reported,Last Date Individual was on a HIDOE Campus,Island,Case Count\n'

#unmergeExcel( filename )
csvdata = headers + convertExcelToCSV( filename )
print(csvdata)

#UNCOMMENT THIS BLOCK TO RESUME USE
#csvdata = headers+openAndCleanExcelFile( filename )
with open('./ListRound1.csv', 'w') as fout:
    fout.write(csvdata)
jsondata = importJsonFromCsv( './ListRound1.csv' )
differentiateOffices( jsondata )
writeJSON( jsondata, 'data.json')
    
#df = pd.read_excel("./List.xlsx")
#df.index = pd.Series(df.index).fillna(method='ffill', axis=0)
#print('read xlsx')
#df.to_csv("./List.csv", sep=",",index = None, header=False)
#print('wrote csv')